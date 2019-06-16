"""
Team Sorter

A program for sorting a group of players into even team based on specified
parameters; namely number of players on a team, avg skill rating of each
team per player, and avg age of each team per player.

Author: Joseph Pruitt (j92pruitt@gmail.com)
"""

from random import randint
import heapq
from openpyxl import load_workbook
from tkinter import *
from tkinter.ttk import *
from tkinter import filedialog


class Player:
    """
    Class holding information about a player for sorting.

    Attributes:
        first_name (string)
        last_name (string)
        rating (float)
    """

    def __init__(self, row_idx, first_name, last_name, rating=0, age=0, request=None):
        self.first_name = first_name
        self.last_name = last_name
        self.rating = rating
        self.row_idx = row_idx
        self.age = age
        if request == None:
            self.request = request
        else:
            self.request = [request]

    def __repr__(self):
        return "{first} {last}".format(
                first=self.first_name, last=self.last_name)

    def add_request(self, player):
        if self.request == None:
            self.request = [player]
        else:
            self.request.append(player)


class Team:
    """
    A list of players, with useful methods and data for
    sorting stored.

    Attributes:
        number (int)
        player_list (list)
        player_count (int)
    """

    def __init__(self, number):
        self.number = number
        self.player_list = []
        self.player_count = 0

    def __repr__(self):
        return "Team {}".format(self.number)

    def __lt__(self, team):
        return self.player_count < team.player_count

    def add_player(self, player):
        self.player_list.append(player)
        self.player_count += 1

    def avg_rating(self):
        total_rating = 0
        for player in self.player_list:
            total_rating += player.rating
        return total_rating/self.player_count

    def avg_age(self):
        total_age = 0
        for player in self.player_list:
            total_age += player.age
        return total_age/self.player_count

    def copy(self):
        team_copy = Team(self.number)
        for player in self.player_list:
            team_copy.add_player(player)
        return team_copy


def sort(playerpool, team_heap):
    """
    This function randomly divides a list of players evenly 
    into a given number of teams.

    Parameters:
        playerpool (list): List of Players to be sorted
        num (int): Number of teams to sort into.

    Returns:
        team_heap (list): List of teams.
    """

    while playerpool:
        idx = randint(0, len(playerpool)-1)
        team = heapq.heappop(team_heap)
        player = playerpool.pop(idx)

        team.add_player(player)
        if player.request:
            for request in player.request:
                team.add_player(request)
                playerpool.remove(request)
        heapq.heappush(team_heap, team)

    return team_heap


def team_sort(playerpool, team_heap, number_of_sorts = 1000):
    """ 
    Executes the sort function number_of_sorts times and returns
    the best result. 
    """

    best_score = 0
    for i in range(number_of_sorts):
        team_heap_copy = [team.copy() for team in team_heap]
        playerpool_copy = playerpool.copy()
        team_list = sort(playerpool_copy, team_heap_copy)

        if i == 0 or sort_score(team_list) < best_score:
            best_teams = team_list
            best_score = sort_score(best_teams)
    return best_teams


def sort_score(team_list):
    """
    Scores a list of team objects, a score of 0 is perfectly
    balanced list of teams. 
    """
    score = 0
    avg_ratings = [team.avg_rating() for team in team_list]
    avg_ages = [team.avg_age() for team in team_list]

    for i in avg_ratings:
        for j in avg_ratings:
            score += (i - j) ** 2
    
    for i in avg_ages:
        for j in avg_ages:
            score += ((i-j) ** 2)/4
    
    player_counts = [team.player_count for team in team_list]
    score *= max(player_counts) - min(player_counts) + 1
    return score


def load_players(worksheet, first_name_col, last_name_col, rating_col, age_col, request_col, num):
    playerpool = []

    team_heap = []
    for i in range(num):
        team_heap.append(Team(i+1))

    for i, row in enumerate(worksheet.values):
        if i > 0:
            playerpool.append(
                Player(i+1, row[first_name_col], row[last_name_col], row[rating_col], row[age_col])
            )
        request_string = row[request_col]
        if request_string:
            if request_string[:9].lower() == "req coach":
                request_team_num = int(request_string[10:])
                team_heap[request_team_num-1].add_player(playerpool[-1])
                print("{} wants to play on team {}".format(playerpool[-1].first_name, request_team_num))
                playerpool.pop()
    
    # Checks row by row for player requests and accomadates them.
    for i, row in enumerate(worksheet.values):
        request_string = row[request_col]
        if request_string:
            if request_string[:10].lower() == "req player":
                request_name = request_string[11:]
                request_name_split = request_name.split()
                first = request_name_split[0].lower()
                last = request_name_split[1].lower()
                for player in playerpool:
                    if player.first_name.lower() == first:
                        if player.last_name.lower() == last:
                            playerpool[i-1].add_request(player)
                            player.add_request(playerpool[i-1])

    heapq.heapify(team_heap)
    
    return playerpool, team_heap


col_num = {}
alphabet_string = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
for i, letter in enumerate(alphabet_string):
    col_num[letter] = i


def text_interface():
    print()
    while True:
        filename = input("What file would you like to sort?: ")
        print()

        try:
            wb = load_workbook(filename, data_only=True)
            break

        except:
            print("Error: Could not load file.\n")

    print("File Loaded. \n")
    print("Available Worksheets are:")

    for i in wb.sheetnames:
        print(i)
    print()

    while True:
        sheetname = input("Which worksheet would you like to sort?:")
        print()

        try:
            ws = wb[sheetname]
            break
        except KeyError:
            print("Error: Incorrect worksheet name.")

    number_of_teams = int(
        input("How many teams would you like for this sort?:")
    )

    playerpool, team_heap = load_players(ws, col_num["D"], col_num["E"], col_num["L"], col_num["O"], col_num["P"], number_of_teams)

    print("There are {} players detected in worksheet".format(len(playerpool)))

    team_list = team_sort(playerpool, team_heap)

    for team in team_list:
        print("Team {}".format(team.number))
        for player in team.player_list:
            print(player)
        print("------------------------------")

def load_file():
    file = filedialog.askopenfilename()
    return file
 
 
def team_sorter_gui(file):
    window = Tk()
    window.title("Team Sorter")
    window.geometry("500x400")
    wb = load_workbook(file, data_only=True)
 
    def load_ws():
        wb.active = wb[ws_selector.get()]
        sort_btn['state'] = 'normal'
        player_count_lbl.configure(text="Worksheet Loaded")
 
    def usr_sort():
        playerpool, team_heap = load_players(wb.active, col_num["D"], col_num["E"], col_num["L"], col_num["O"], col_num["P"], int(team_count_select.get()))
        team_list = team_sort(playerpool, team_heap)
        for team in team_list:
            for player in team.player_list:
                cell = "A" + str(player.row_idx)
                wb.active[cell] = team.number
        wb.save(file)
        feedback_lbl.configure(text="Sort Complete. Sort Score: {}".format(sort_score(team_list)))
 
    ws_lbl = Label(window, text="Available Worksheets:")
    ws_lbl.grid(row=0, column=0)
    ws_selector = Combobox(window)
    ws_selector['values'] = tuple(wb.sheetnames)
    ws_selector.grid(row=0, column=1)
    load_players_btn = Button(window, text="Select", command=load_ws)
    load_players_btn.grid(row=0, column=2)
 
    player_count_lbl = Label(window, text="No worksheet loaded")
    player_count_lbl.grid(row=1, column=0)
 
    team_count_lbl = Label(window, text="Number of teams:")
    team_count_lbl.grid(row=2, column=0)
    team_count_select = Combobox(window)
    team_count_select['values'] = (1, 2, 3, 4, 5, 6)
    team_count_select.current(0)
    team_count_select.grid(row=2, column=1)
 
    sort_btn = Button(window, text="Sort", command=usr_sort, state=DISABLED)
    sort_btn.grid(row=2, column=2)
 
    feedback_lbl = Label(window, text="")
    feedback_lbl.grid(row=3, column=4)
 
    window.mainloop()

text_interface()
#input_file = load_file()
#team_sorter_gui(input_file)
    